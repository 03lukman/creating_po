import openpyxl
from openpyxl.styles import Alignment, PatternFill
import os
from datetime import datetime


def export_excel(data, columns, save_dir):
    os.makedirs(save_dir, exist_ok=True)

    columns.append("HARGA")
    columns.append("TOTAL")

    filename = f"rekap_so_final_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(save_dir, filename)

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.append(columns)

    fill_red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    for row in data:
        sheet.append(row)

    for r in sheet.iter_rows(min_row=2):
        r[0].number_format = 'dd/mm/yy'
        r[4].number_format = '#,##0'

        harga_cell = r[-2]
        total_cell = r[-1]

        if isinstance(harga_cell.value, (int, float)):
            harga_cell.number_format = '#,##0'
        else:
            harga_cell.value = "NOT FOUND"
            harga_cell.fill = fill_red

        if isinstance(total_cell.value, (int, float)):
            total_cell.number_format = '#,##0'
        else:
            total_cell.value = ""
            total_cell.fill = fill_red

        for cell in r:
            cell.alignment = Alignment(vertical='top')

    wb.save(output_path)

    return output_path